"""
Export service for generating various output formats from extracted transactions.
Supports CSV, Excel, QBO, OFX, JSON, MT940, and CAMT.053.
"""

from __future__ import annotations

import csv
import io
import json
import logging
from datetime import datetime, timezone
from decimal import Decimal
from typing import Any, Optional

from core.config import get_settings

settings = get_settings()
logger = logging.getLogger(__name__)


class ExportService:
    """Service for exporting transactions to various formats."""

    SUPPORTED_FORMATS = {
        "csv": "text/csv",
        "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "json": "application/json",
        "qbo": "application/vnd.intuit.qbo",
        "ofx": "application/x-ofx",
        "mt940": "text/plain",
        "camt053": "application/xml",
    }

    @classmethod
    def get_content_type(cls, format: str) -> str:
        return cls.SUPPORTED_FORMATS.get(format, "application/octet-stream")

    @classmethod
    def generate_csv(
        cls,
        transactions: list[dict],
        metadata: dict = None,
        include_headers: bool = True,
    ) -> bytes:
        """Generate CSV export."""
        output = io.StringIO()
        writer = csv.writer(output)

        if include_headers and metadata:
            writer.writerow(["Statementwise Export"])
            writer.writerow(
                [
                    "Bank:",
                    metadata.get("bank_name", ""),
                    "Account:",
                    metadata.get("account_number", ""),
                ]
            )
            writer.writerow(
                [
                    "Period:",
                    f"{metadata.get('statement_period', {}).get('start_date', '')} to {metadata.get('statement_period', {}).get('end_date', '')}",
                ]
            )
            writer.writerow([])

        # Transaction headers
        writer.writerow(
            ["Date", "Description", "Reference", "Category", "Debit", "Credit", "Amount", "Currency", "Balance"]
        )

        for tx in transactions:
            writer.writerow(
                [
                    tx.get("date", ""),
                    tx.get("description", ""),
                    tx.get("reference", ""),
                    tx.get("category", ""),
                    f"{tx.get('debit', 0) or 0:.2f}" if tx.get("debit") else "",
                    f"{tx.get('credit', 0) or 0:.2f}" if tx.get("credit") else "",
                    f"{tx.get('amount', 0):.2f}",
                    tx.get("currency", "USD"),
                    f"{tx.get('balance', 0) or 0:.2f}" if tx.get("balance") else "",
                ]
            )

        return output.getvalue().encode("utf-8")

    @classmethod
    def generate_json(
        cls,
        transactions: list[dict],
        metadata: dict = None,
        opening_balance: dict = None,
        closing_balance: dict = None,
        summary: dict = None,
        reconciliation: dict = None,
    ) -> bytes:
        """Generate comprehensive JSON export."""
        data = {
            "export_metadata": {
                "exported_at": datetime.now(timezone.utc).isoformat(),
                "format_version": "1.0",
                "total_transactions": len(transactions),
            },
            "statement_metadata": metadata or {},
            "opening_balance": opening_balance or {},
            "closing_balance": closing_balance or {},
            "summary": summary or {},
            "reconciliation": reconciliation or {},
            "transactions": transactions,
        }
        return json.dumps(data, indent=2, default=str).encode("utf-8")

    @classmethod
    def generate_excel(
        cls,
        transactions: list[dict],
        metadata: dict = None,
        opening_balance: dict = None,
        closing_balance: dict = None,
        summary: dict = None,
    ) -> bytes:
        """Generate Excel XLSX export."""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error("openpyxl required for Excel export")
            raise ImportError("Install openpyxl: pip install openpyxl")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Metadata section
        if metadata:
            ws.append(["Statementwise Bank Statement Export"])
            ws.append(["Bank", metadata.get("bank_name", "")])
            ws.append(["Account Holder", metadata.get("account_holder", "")])
            ws.append(["Account Number", metadata.get("account_number", "")])
            ws.append(
                [
                    "Period",
                    f"{metadata.get('statement_period', {}).get('start_date', '')} to {metadata.get('statement_period', {}).get('end_date', '')}",
                ]
            )
            if opening_balance:
                ws.append(["Opening Balance", f"{opening_balance.get('amount', 0):.2f}"])
            if closing_balance:
                ws.append(["Closing Balance", f"{closing_balance.get('amount', 0):.2f}"])
            ws.append([])

        # Headers
        headers = ["Date", "Description", "Reference", "Category", "Debit", "Credit", "Amount", "Currency", "Balance"]
        ws.append(headers)

        header_row = ws.max_row
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Transaction data
        for tx in transactions:
            ws.append(
                [
                    tx.get("date", ""),
                    tx.get("description", ""),
                    tx.get("reference", ""),
                    tx.get("category", ""),
                    tx.get("debit") if tx.get("debit") else 0,
                    tx.get("credit") if tx.get("credit") else 0,
                    tx.get("amount", 0),
                    tx.get("currency", "USD"),
                    tx.get("balance") if tx.get("balance") else "",
                ]
            )

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            max_length = 0
            column = get_column_letter(col)
            for row in ws.iter_rows(min_col=col, max_col=col):
                for cell in row:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
            ws.column_dimensions[column].width = min(max_length + 2, 60)

        # Summary sheet
        if summary:
            ws_summary = wb.create_sheet(title="Summary")
            ws_summary.append(["Statement Summary"])
            ws_summary.append([])
            ws_summary.append(["Total Credits", summary.get("total_credits", 0)])
            ws_summary.append(["Total Debits", summary.get("total_debits", 0)])
            ws_summary.append(["Net Change", summary.get("total_credits", 0) - summary.get("total_debits", 0)])
            ws_summary.append(["Transaction Count", summary.get("transaction_count", 0)])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @classmethod
    def generate_qbo(
        cls,
        transactions: list[dict],
        metadata: dict = None,
        opening_balance: dict = None,
        closing_balance: dict = None,
    ) -> bytes:
        """Generate QuickBooks QBO (OFX 2.1.1) export."""
        lines = [
            'OFXHEADER:100',
            'DATA:OFXSGML',
            'VERSION:102',
            'SECURITY:NONE',
            'ENCODING:USASCII',
            'CHARSET:1252',
            'COMPRESSION:NONE',
            'OLDFILEUID:NONE',
            'NEWFILEUID:NONE',
            '',
            '<OFX>',
            '  <SIGNONMSGSRSV1>',
            '    <SONRS>',
            '      <STATUS>',
            '        <CODE>0</CODE>',
            '        <SEVERITY>INFO</SEVERITY>',
            '      </STATUS>',
            f'      <DTSERVER>{datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")}</DTSERVER>',
            '      <LANGUAGE>ENG</LANGUAGE>',
            '    </SONRS>',
            '  </SIGNONMSGSRSV1>',
            '  <BANKMSGSRSV1>',
            '    <STMTTRNRS>',
            '      <TRNUID>1001</TRNUID>',
            '      <STATUS>',
            '        <CODE>0</CODE>',
            '        <SEVERITY>INFO</SEVERITY>',
            '      </STATUS>',
            '      <STMTRS>',
            '        <CURDEF>USD</CURDEF>',
            '        <BANKACCTFROM>',
            f'          <BANKID>{metadata.get("bank_name", "")}</BANKID>',
            f'          <ACCTID>{metadata.get("account_number", "").replace("*", "")}</ACCTID>',
            '          <ACCTTYPE>CHECKING</ACCTTYPE>',
            '        </BANKACCTFROM>',
            '        <BANKTRANLIST>',
        ]

        if metadata and metadata.get("statement_period"):
            start_date = metadata["statement_period"].get("start_date", "").replace("-", "")
            end_date = metadata["statement_period"].get("end_date", "").replace("-", "")
            lines.append(f'          <DTSTART>{start_date}</DTSTART>')
            lines.append(f'          <DTEND>{end_date}</DTEND>')

        for i, tx in enumerate(transactions):
            date = tx.get("date", "").replace("-", "")
            amount = tx.get("amount", 0)
            if tx.get("debit"):
                amount = -abs(amount)
            else:
                amount = abs(amount)

            lines.extend(
                [
                    '          <STMTTRN>',
                    f'            <TRNTYPE>{("DEBIT" if amount < 0 else "CREDIT")}</TRNTYPE>',
                    f'            <DTPOSTED>{date}</DTPOSTED>',
                    f'            <TRNAMT>{amount:.2f}</TRNAMT>',
                    f'            <FITID>TXN{str(i + 1).zfill(6)}</FITID>',
                    f'            <NAME>{cls._escape_xml(tx.get("description", ""))}</NAME>',
                    f'            <REFNUM>{cls._escape_xml(tx.get("reference", ""))}</REFNUM>',
                    '          </STMTTRN>',
                ]
            )

        lines.extend(
            [
                '        </BANKTRANLIST>',
                '        <LEDGERBAL>',
            ]
        )

        if closing_balance:
            amount = closing_balance.get("amount", 0)
            date = closing_balance.get("date", "").replace("-", "")
            lines.append(f'          <BALAMT>{amount:.2f}</BALAMT>')
            lines.append(f'          <DTASOF>{date}</DTASOF>')

        lines.extend(
            [
                '        </LEDGERBAL>',
                '      </STMTRS>',
                '    </STMTTRNRS>',
                '  </BANKMSGSRSV1>',
                '</OFX>',
            ]
        )

        return "\n".join(lines).encode("utf-8")

    @classmethod
    def generate_ofx(
        cls,
        transactions: list[dict],
        metadata: dict = None,
        opening_balance: dict = None,
        closing_balance: dict = None,
    ) -> bytes:
        """Generate OFX (Open Financial Exchange) export - same as QBO format."""
        return cls.generate_qbo(transactions, metadata, opening_balance, closing_balance)

    @classmethod
    def generate_mt940(
        cls,
        transactions: list[dict],
        metadata: dict = None,
        opening_balance: dict = None,
        closing_balance: dict = None,
    ) -> bytes:
        """Generate MT940 format export."""
        lines = []
        account = metadata.get("account_number", "").replace("*", "") if metadata else ""
        bank = metadata.get("bank_name", "") if metadata else ""

        # Header
        lines.append(f":20:STMTWISE-{datetime.now(timezone.utc).strftime('%Y%m%d')}")
        lines.append(f":25:{account}")
        lines.append(f":28C:1/1")

        # Opening balance
        if opening_balance:
            amount = opening_balance.get("amount", 0)
            date = opening_balance.get("date", "").replace("-", "")
            sign = "C" if amount >= 0 else "D"
            lines.append(f":60F:{sign}{date}USD{abs(amount):.2f}")

        # Transactions
        for tx in transactions:
            date = tx.get("date", "").replace("-", "")[2:] if tx.get("date") else ""
            amount = tx.get("amount", 0)
            sign = "D" if tx.get("debit") else "C"
            desc = tx.get("description", "")[:35]
            ref = (tx.get("reference", "") or "")[:16]

            lines.append(f":61:{date}{sign}{abs(amount):.2f}NMSC{ref}")
            lines.append(f":86:{desc}")

        # Closing balance
        if closing_balance:
            amount = closing_balance.get("amount", 0)
            date = closing_balance.get("date", "").replace("-", "")
            sign = "C" if amount >= 0 else "D"
            lines.append(f":62F:{sign}{date}USD{abs(amount):.2f}")

        lines.append("-")
        return "\n".join(lines).encode("utf-8")

    @classmethod
    def generate_camt053(
        cls,
        transactions: list[dict],
        metadata: dict = None,
        opening_balance: dict = None,
        closing_balance: dict = None,
    ) -> bytes:
        """Generate CAMT.053 (ISO 20022) XML export."""
        account = metadata.get("account_number", "").replace("*", "") if metadata else ""
        bank = metadata.get("bank_name", "") if metadata else ""
        iban = metadata.get("iban", account)
        currency = (transactions[0].get("currency", "USD") if transactions else "USD")

        lines = [
            '<?xml version="1.0" encoding="UTF-8"?>',
            '<Document xmlns="urn:iso:std:iso:20022:tech:xsd:camt.053.001.02">',
            '  <BkToCstmrStmt>',
            f'    <GrpHdr>',
            f'      <MsgId>STMTWISE-{datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")}</MsgId>',
            f'      <CreDtTm>{datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S")}</CreDtTm>',
            f'    </GrpHdr>',
            f'    <Stmt>',
            f'      <Id>STMT-1</Id>',
            f'      <ElctrncSeqNb>1</ElctrncSeqNb>',
        ]

        if metadata and metadata.get("statement_period"):
            start_date = metadata["statement_period"].get("start_date", "")
            end_date = metadata["statement_period"].get("end_date", "")
            lines.append(f'      <FrToDt>')
            lines.append(f'        <FrDtTm>{start_date}T00:00:00</FrDtTm>')
            lines.append(f'        <ToDtTm>{end_date}T23:59:59</ToDtTm>')
            lines.append(f'      </FrToDt>')

        lines.extend(
            [
                f'      <Acct>',
                f'        <Id>',
                f'          <IBAN>{iban}</IBAN>',
                f'        </Id>',
                f'        <Ccy>{currency}</Ccy>',
                f'      </Acct>',
            ]
        )

        # Opening balance
        if opening_balance:
            amount = opening_balance.get("amount", 0)
            lines.extend(
                [
                    f'      <Bal>',
                    f'        <Tp><CdOrPrtry><Cd>OPBD</Cd></CdOrPrtry></Tp>',
                    f'        <Amt Ccy="{currency}">{amount:.2f}</Amt>',
                    f'        <CdtDbtInd>{("CRDT" if amount >= 0 else "DBIT")}</CdtDbtInd>',
                    f'        <Dt><Dt>{opening_balance.get("date", "")}</Dt></Dt>',
                    f'      </Bal>',
                ]
            )

        # Transactions
        lines.append(f'      <Ntry>')
        for tx in transactions:
            amount = tx.get("amount", 0)
            is_credit = not tx.get("debit")
            lines.extend(
                [
                    f'        <Ntry>',
                    f'          <Amt Ccy="{currency}">{amount:.2f}</Amt>',
                    f'          <CdtDbtInd>{("CRDT" if is_credit else "DBIT")}</CdtDbtInd>',
                    f'          <Sts>BOOK</Sts>',
                    f'          <BookgDt><Dt>{tx.get("date", "")}</Dt></BookgDt>',
                    f'          <ValDt><Dt>{tx.get("date", "")}</Dt></ValDt>',
                    f'          <AcctSvcrRef>{cls._escape_xml(tx.get("reference", ""))}</AcctSvcrRef>',
                    f'          <NtryDtls>',
                    f'            <TxDtls>',
                    f'              <RmtInf>',
                    f'                <Ustrd>{cls._escape_xml(tx.get("description", ""))}</Ustrd>',
                    f'              </RmtInf>',
                    f'            </TxDtls>',
                    f'          </NtryDtls>',
                    f'        </Ntry>',
                ]
            )

        # Closing balance
        if closing_balance:
            amount = closing_balance.get("amount", 0)
            lines.extend(
                [
                    f'      <Bal>',
                    f'        <Tp><CdOrPrtry><Cd>CLBD</Cd></CdOrPrtry></Tp>',
                    f'        <Amt Ccy="{currency}">{amount:.2f}</Amt>',
                    f'        <CdtDbtInd>{("CRDT" if amount >= 0 else "DBIT")}</CdtDbtInd>',
                    f'        <Dt><Dt>{closing_balance.get("date", "")}</Dt></Dt>',
                    f'      </Bal>',
                ]
            )

        lines.extend(
            [
                f'    </Stmt>',
                f'  </BkToCstmrStmt>',
                f'</Document>',
            ]
        )

        return "\n".join(lines).encode("utf-8")

    @classmethod
    def generate_export(
        cls,
        format: str,
        transactions: list[dict],
        metadata: dict = None,
        opening_balance: dict = None,
        closing_balance: dict = None,
        summary: dict = None,
        reconciliation: dict = None,
    ) -> tuple[bytes, str, str]:
        """
        Generate export in specified format.

        Returns:
            Tuple of (content_bytes, content_type, filename)
        """
        format = format.lower()
        if format not in cls.SUPPORTED_FORMATS:
            raise ValueError(f"Unsupported format: {format}")

        content_type = cls.SUPPORTED_FORMATS[format]
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        bank = (metadata.get("bank_name", "statement") if metadata else "statement")
        bank = bank.lower().replace(" ", "_")
        filename = f"{bank}_export_{timestamp}.{format}"

        generators = {
            "csv": lambda: cls.generate_csv(transactions, metadata),
            "xlsx": lambda: cls.generate_excel(transactions, metadata, opening_balance, closing_balance, summary),
            "json": lambda: cls.generate_json(transactions, metadata, opening_balance, closing_balance, summary, reconciliation),
            "qbo": lambda: cls.generate_qbo(transactions, metadata, opening_balance, closing_balance),
            "ofx": lambda: cls.generate_ofx(transactions, metadata, opening_balance, closing_balance),
            "mt940": lambda: cls.generate_mt940(transactions, metadata, opening_balance, closing_balance),
            "camt053": lambda: cls.generate_camt053(transactions, metadata, opening_balance, closing_balance),
        }

        content = generators[format]()
        return content, content_type, filename

    @staticmethod
    def _escape_xml(text: str) -> str:
        """Escape XML special characters."""
        if not text:
            return ""
        return (
            text.replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', "&quot;")
            .replace("'", "&apos;")
        )
